import openpyxl
from openpyxl.styles import Alignment
import json

def create_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比结果"

    # 写入表头
    headers = ["学院", "原始链接", "核心摘要", "共同点", "差异点", "独特亮点" ]
    ws.append(headers)

    # 写入数据
    for row in data:
        ws.append(row)

    # 设置自动换行和列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                max_length = max(max_length, len(str(cell.value)) // 2)  # 粗略估算
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # 冻结首行
    ws.freeze_panes = "A2"
    wb.save(output_path)

def jsonl_to_xlsx(jsonl_path, xlsx_path, other_msgs:[str]):
    """将 JSONL 文件转换为 Excel (.xlsx)，并对长文本列启用自动换行"""
    import json
    import sys
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font
    except ImportError:
        print("错误: 需要安装 openpyxl 库。请运行: pip install openpyxl")
        return

    # 读取所有行
    rows = []
    fieldnames = None
    with open(jsonl_path, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                if fieldnames is None:
                    fieldnames = list(row.keys())
                # 确保字段顺序一致
                ordered_row = [row.get(f, '') for f in fieldnames]
                print(f"{line_num}: {ordered_row}")

                rows.append(ordered_row)
            except json.JSONDecodeError as e:
                print(f"警告: 第 {line_num} 行 JSON 解析失败: {e}", file=sys.stderr)

    if not rows:
        print("没有有效数据", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比结果"

    # 写入表头
    for col_idx, header in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)          # 粗体
        cell.alignment = Alignment(horizontal='center', vertical='center')  # 居中

    # 写入数据
    for row in rows:
        ws.append(row)

    for msg in other_msgs:
        ws.append([msg])
        ws.merge_cells(start_row=ws.max_row, start_column=1,end_row=ws.max_row, end_column=ws.max_column)

    # 设置自动换行和列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.value:
                # 粗略估计列宽：取最长字符数的一半（中文字符占2）
                val_len = len(str(cell.value))
                # 中文字符比例估算，简单处理：取字符数
                max_len = max(max_len, val_len)
        # 限制最大列宽 150，最小 30
        adjusted_width = min(max(max_len // 2, 30), 150)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = 'A2'

    wb.save(xlsx_path)
    print(f"✓ Excel 已保存至: {xlsx_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="将 JSONL 文件转换为 Excel (.xlsx)，并对长文本列启用自动换行")
    parser.add_argument("jsonl_path", help="输入 JSONL 文件路径")
    parser.add_argument("xlsx_path", help="输出 Excel 文件路径")
    parser.add_argument("--other_msgs", nargs='*', default=[], help="其他消息，作为独立行添加到表格末尾")
    args = parser.parse_args()

    jsonl_to_xlsx(args.jsonl_path, args.xlsx_path, args.other_msgs)
